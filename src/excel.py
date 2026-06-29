import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "output", "data.xlsx")
EXCEL_PATH = os.path.normpath(EXCEL_PATH)
HEADERS = ["SR.NO", "NAME", "VILLAGE", "MOBILE.NO", "AADHAAR.NO", "AP.DATE", "CENTER"]

# Column widths (in Excel units)
COL_WIDTHS = {
    "A": 8,   # SR.NO
    "B": 28,  # NAME
    "C": 20,  # VILLAGE
    "D": 15,  # MOBILE.NO
    "E": 16,  # AADHAAR.NO
    "F": 13,  # AP.DATE
    "G": 16,  # CENTER
}

# Starting SR.NO — can be overridden by the UI
_start_sr_no = None


def set_start_sr_no(value: int):
    """Set the starting SR.NO for this session."""
    global _start_sr_no
    _start_sr_no = value


def _thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_header(ws):
    """Apply header styling: bold, font 10, background, borders, alignment."""
    header_fill = PatternFill("solid", fgColor="4472C4")  # blue header
    for col, cell in enumerate(ws[1], start=1):
        cell.font      = Font(bold=True, size=10, color="FFFFFF")
        cell.fill      = header_fill
        cell.border    = _thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 20


def _style_data_row(ws, row_num: int):
    """Apply font size 10, borders, and alignment to a data row."""
    for cell in ws[row_num]:
        cell.font      = Font(size=10)
        cell.border    = _thin_border()
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    # SR.NO center aligned
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_num].height = 16


def _apply_col_widths(ws):
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width


def get_or_create_workbook():
    """Load existing workbook or create a new one with styled headers."""
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Records"
        ws.append(HEADERS)
        _style_header(ws)
        _apply_col_widths(ws)
        # Freeze the header row
        ws.freeze_panes = "A2"
        wb.save(EXCEL_PATH)
    return wb, ws


def get_next_sr_no() -> int:
    """Return the next SR.NO to be assigned."""
    _, ws = get_or_create_workbook()
    data_rows = ws.max_row - 1  # subtract header row
    if _start_sr_no is not None:
        return _start_sr_no + data_rows
    else:
        return data_rows + 1  # start from 1


def save_record(fields: dict) -> int:
    """Append a new record to Excel with styling. Returns the SR.NO assigned."""
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    wb, ws = get_or_create_workbook()

    sr_no = get_next_sr_no()
    row = [
        sr_no,
        fields.get("NAME", ""),
        fields.get("VILLAGE", ""),
        fields.get("MOBILE.NO", ""),
        fields.get("AADHAAR.NO", ""),
        fields.get("AP.DATE", ""),
        fields.get("CENTER", ""),
    ]
    ws.append(row)
    _style_data_row(ws, ws.max_row)
    wb.save(EXCEL_PATH)
    return sr_no
