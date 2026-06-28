import openpyxl
import os

EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "output", "data.xlsx")
EXCEL_PATH = os.path.normpath(EXCEL_PATH)
HEADERS = ["SR.NO", "NAME", "VILLAGE", "MOBILE.NO", "AADHAAR.NO", "AP.DATE", "CENTER"]


def get_or_create_workbook():
    """Load existing workbook or create a new one with headers."""
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Records"
        ws.append(HEADERS)
        # Style the header row
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        wb.save(EXCEL_PATH)
    return wb, ws


def get_next_sr_no():
    """Return the next SR.NO based on existing rows."""
    _, ws = get_or_create_workbook()
    # Subtract 1 for header row
    return ws.max_row


def save_record(fields: dict) -> int:
    """
    Append a new record to the Excel file.
    fields: dict with keys NAME, VILLAGE, MOBILE.NO, AADHAAR.NO, AP.DATE, CENTER
    Returns the SR.NO assigned.
    """
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    wb, ws = get_or_create_workbook()

    sr_no = ws.max_row  # header is row 1, so first data row gives sr_no = 1
    row = [
        sr_no,
        fields.get("NAME", ""),
        fields.get("VILLAGE", ""),
        fields.get("MOBILE.NO", ""),
        fields.get("AADHAAR.NO", ""),
        fields.get("AP.DATE", ""),
        fields.get("CENTER", ""),
    ]
    ws.append(row)
    wb.save(EXCEL_PATH)
    return sr_no